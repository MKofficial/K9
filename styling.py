from my_utils import *
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
import openpyxl as opx

heading = ['Pořadí', 'Jméno a příjmení', 'Klub', 'Distance', 'Kategorie', '1. závod', '2. závod', '3. závod',
           '4. závod', '5. závod', '6. závod', '7. závod', '8. závod', '9. závod', '10. závod', '11. závod',
           '12. závod', 'Body']


def style_and_save(excel_workbook, excel_workbook_name, excel_workbook_sheet, styling_function):
    """
    :param excel_workbook: Workbook from openpyxl library
    :param excel_workbook_name: Name of the the certain workbook
    :param styling_function: Function that will style the table
    :param excel_workbook_sheet: Workbook sheet from openpyxl library
    :return: None

    This function apply the certain styling function to the current table and save it
    """

    styling_function(excel_workbook_sheet)
    excel_workbook.save(excel_workbook_name)


# there has to be defined workbook and sheet which is passed as an argument to excel_file parameter
# WARNING you have to save the sheet manually in the code
def preposition(excel_sheet) -> None:
    """
    :param excel_sheet: Excel sheet to be styled
    :return: None

    These change will be visible after saving that workbook with openpyxl
    WARNING this function does not support saving
    """
    try:
        excel_sheet.insert_rows(0)
        excel_sheet.insert_cols(0)
        excel_sheet.append(heading)
        excel_sheet.move_range('A' + str(excel_sheet.max_row) + ':R' + str(excel_sheet.max_row),
                               rows=-excel_sheet.max_row + 1)
    except FileNotFoundError:
        raise_error(excel_sheet, problem_message[0])
    except InvalidFileException:
        raise_error(excel_sheet, problem_message[2])


def category_and_position(workbook_name, array: list, race_number: int) -> None:
    """
    :param workbook_name: Excel workbook name
    :param array: Array containing people and their points
    :param race_number: An integer defining the race number to make "Body" in right position
    :return: None
    """
    workbook = opx.Workbook()
    sheet = workbook.active

    headline = ['Pořadí', 'Jméno a příjmení', 'Klub', '1. závod', '2. závod', '3. závod', '4. závod',
                '5. závod', '6. závod', '7. závod', '8. závod', '9. závod', '10. závod', '11. závod', '12. závod',
                'Body']
    for param in dist_and_cat:
        position = 0
        temp = []

        for person in array:
            if person[2] == param[0] and person[3] == param[1]:
                # warning...deleting position will change position of other values!!!
                # deleting category and route
                del person[3]
                del person[2]

                # append the person to temporary array
                temp.append(person)

        # sort persons in temporary array with the last value of each field - sum of points
        temp.sort(key=lambda x: x[-1], reverse=True)

        # to check if there are some people having same sum of points
        last_pos = None
        for i in temp:
            if i[-1] == 0:
                i.insert(0, "-")
            elif last_pos != i[-1]:
                position += 1
                last_pos = i[-1]
                i.insert(0, position)
            else:
                i.insert(0, position)
            # insert None to manage the sum of points into the right column
            for j in range(12 - race_number):
                i.insert(-1, None)

        temp.insert(0, headline)
        temp.insert(0, (param[0], param[1]))

        for i in temp:
            sheet.append(i)

        workbook.save(workbook_name)


def gui_elimination_of_points(workbook_name: str) -> None:
    """
    :param workbook_name: Excel workbook name
    :return: None
    """

    workbook = opx.load_workbook(workbook_name)
    sheet = workbook.active

    for i in sheet['A']:
        if sheet[i.row][3].value is None or sheet[i.row][3].value == '1. závod':
            continue
        else:
            none_counter = 0
            position = [[3, sheet[i.row][3].value], [4, sheet[i.row][4].value], [5, sheet[i.row][5].value],
                        [6, sheet[i.row][6].value], [7, sheet[i.row][7].value], [8, sheet[i.row][8].value],
                        [9, sheet[i.row][9].value], [10, sheet[i.row][10].value], [11, sheet[i.row][11].value],
                        [12, sheet[i.row][12].value], [13, sheet[i.row][13].value], [14, sheet[i.row][14].value]]

            # TODO: adjust ->
            for elements in position:
                if elements[1] is None or elements[1] == '-' or elements[1] in unfinished:
                    none_counter += 1
            if none_counter < 4:
                temp = []
                for elements in position:
                    if elements[1] != '-' and elements[1] is not None and elements[1] not in unfinished:
                        temp.append(elements)
                temp.sort(key=lambda x: x[-1])
                cells = [temp[k] for k in range(4 - none_counter)]
                for j in cells:
                    sheet[i.row][j[0]].font = Font(color='c0c0c0', italic=True, strikethrough=True)

    # setup font
    for i in sheet:
        for j in i:
            j.font = Font(name='Arial', size=10)

    workbook.save(workbook_name)


def color(workbook_name: str) -> None:
    """
    :param workbook_name: Excel workbook name
    :return: None
    """

    workbook = opx.load_workbook(workbook_name)
    sheet = workbook.active

    # styling headline of each category
    for i in sheet['A']:
        if i.value == 'Dlouhý okruh - 9 km' or i.value == 'Krátký okruh - 4,5 km':
            for j in sheet[i.row]:
                j.fill = PatternFill(start_color='c0c0c0', end_color='c0c0c0', fill_type='solid')
                if j == "<Cell 'Sheet'.P" + str(sheet[i.row]) + ">":
                    break

    workbook.save(workbook_name)

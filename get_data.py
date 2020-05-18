import openpyxl as opx
from openpyxl.utils.exceptions import InvalidFileException
from my_utils import *

unfinished = ["DNF", "DSQ", "DNS"]


def add_points(people: list = [], multiply: bool = False) -> list:
    """
    :param people: List with people to be added points as their last element
    :param multiply: This multiply the points. Usage only if the K9 is on the sixth race
    :return: List with people containing points now
    """
    points = 11
    if multiply:
        points *= 1.5

    for person in people:
        if person[2] in unfinished:
            for i in unfinished:
                if person[2] == i:
                    person.append(i)
                    del person[2]
        if points == 10:
            points -= 1
        if points <= 0:
            person.append(0)
            del person[2]
        else:
            person.append(points)
            points -= 1
            del person[2]

    return people


def get_file_arr(excel_file: str = 'kuneticka_devitka.xlsx', distance: str = 'Dlouhý okruh - 9 km',
                 category: str = 'Muži 18 - 29 let', sixth_race: bool = False) -> list:
    """
    :param sixth_race: If it is sixth race, then multiply the points in add_points() by 1.5
    :param excel_file: Excel file with the newest race's results
    :param category: Person's category
    :param distance: Person's distance
    :return: List that contains persons from specific category and distance
    """
    try:
        # try to open new workbook
        wb = opx.load_workbook(excel_file)
        ws = wb.active
    except InvalidFileException:
        raise_error(excel_file, problem_message[2])
    except FileNotFoundError:
        raise_error(excel_file, problem_message[0])
    else:

        people = [[ws[person.row][1].value, ws[person.row][4].value, ws[person.row][11].value,
                   ws[person.row][5].value, ws[person.row][6].value] for person in ws['F']
                  if person.value == distance and ws[person.row][6].value == category
                  if ws[person.row][1].value.lower() != 'jméno a příjmení']

    # add points to participates in each category
    add_points(people, sixth_race)

    return people


def get_complete_file_arr(excel_file: str = 'kuneticka_devitka.xlsx', sixth_race: bool = False) -> list:
    """
    :param sixth_race: If it is sixth race then set last param in get_file_arr() to True
    :param excel_file: File with the race results
    :return: All participants of that race
    """

    if sixth_race:
        people = [cat_people for param in dist_and_cat for cat_people in
                  get_file_arr(excel_file, param[0], param[1], sixth_race=True)]
    elif sixth_race is False:
        people = [cat_people for param in dist_and_cat for cat_people in
                  get_file_arr(excel_file, param[0], param[1], sixth_race=False)]
    else:
        raise_error(message=problem_message[3])

    return people


def get_final_arr(excel_file: str = 'final_table.xlsx') -> list:
    """
    :param excel_file: Excel file with the complete results
    :return: List of all elements in there
    """
    try:
        wb = opx.load_workbook(excel_file)
        ws = wb.active
    except InvalidFileException:
        raise_error(excel_file, problem_message[2])
    except FileNotFoundError:
        raise_error(excel_file, problem_message[0])
    else:
        if len(ws.column_dimensions) == 0:
            people = []
        else:
            people = [[person.value for person in element] for element in ws]

    return people
